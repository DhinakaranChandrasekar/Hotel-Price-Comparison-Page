import time
import datetime
import re
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from rapidfuzz import process as fuzz_process, fuzz
from concurrent.futures import ThreadPoolExecutor

# Function to extract bedroom count from room name
def extract_bedroom_count(room_name):
    match = re.search(r'(\d+)\s*Bedroom', room_name, re.IGNORECASE)
    return int(match.group(1)) if match else 1

# Function to scrape IOL data
def scrape_iol(iol_hotel_url):
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("window-size=1920,1080")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.get("https://b2b.iol-x.com/")
        print("IOL Page loaded.")

        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "email")))
        email_field = driver.find_element(By.ID, "email")
        password_field = driver.find_element(By.ID, "password")

        email_field.send_keys("dhinakaran.c+edgeb2btravelagent@iol.world")
        password_field.send_keys("Supermen61977@")

        login_button = driver.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        print("IOL Logged in.")

        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "DestinationInput")))
        
        driver.get(iol_hotel_url)
        print("Navigated to IOL hotel URL.")
        
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, 'h1')))
        print("IOL hotel page loaded successfully.")

        try:
            view_more_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "View all room types")]'))
            )
            driver.execute_script("arguments[0].click();", view_more_button)
            time.sleep(5)
        except:
            pass

        iol_room_data = []
        try:
            hotel_name_iol = driver.find_element(By.TAG_NAME, 'h1').text

            date_range_element = driver.find_element(By.CSS_SELECTOR, '.hotel_trip_date_range')
            date_range = date_range_element.text.split(' - ')[0].strip()
            start_date = datetime.datetime.strptime(date_range, "%a, %b %d, %Y").strftime("%d-%b-%y")

            room_elements = driver.find_elements(By.CSS_SELECTOR, '.hotel_room_item')
            for room_element in room_elements:
                room_type = room_element.find_element(By.CSS_SELECTOR, 'h3.hotel_room_type_name').text
                print(f"IOL Room Type: {room_type}")

                meal_plan_element = room_element.find_element(By.CSS_SELECTOR, '.room_mealplan_details')
                meal_plan_text = meal_plan_element.text if meal_plan_element else 'N/A'
                meal_plan = meal_plan_text.split('ContractToken')[0].strip()

                cancellation_element = room_element.find_element(By.CSS_SELECTOR, "div.hotel_room_selection.hotel_room_cancellation_policy span, div.hotel_room_selection.hotel_room_cancellation_policy p")

                cancellation_text = cancellation_element.text if cancellation_element else 'N/A'
                cancellation = 'Non-refundable' if 'Non-refundable' in cancellation_text else 'Free cancellation'

                price_element = room_element.find_element(By.CSS_SELECTOR, '.hotel_room_discounted_price')
                price_text = price_element.text if price_element else 'N/A'
                price = price_text.split(' ')[-1]

                bedrooms = extract_bedroom_count(room_type)

                iol_room_data.append((room_type, meal_plan, cancellation, price, bedrooms))
        except Exception as e:
            print("Error during IOL scraping:", e)
        
        return iol_room_data, start_date, hotel_name_iol

    finally:
        driver.quit()
        print("IOL driver closed.")

# Function to scrape RateHawk data
def scrape_ratehawk(ratehawk_hotel_url):
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--headless")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    try:
        driver.get(ratehawk_hotel_url)
        
        ratehawk_cookies = [
            {"name": "uid", "value": "TfTb5WYnpUFawFYpHpPPAg==", "domain": ".ratehawk.com"},
            {"name": "riskified_sid", "value": "f9f5b754-5e72-4a58-9482-f67d197d8259", "domain": ".ratehawk.com"},
            {"name": "_gcl_au", "value": "1.1.932512046.1713874242", "domain": ".ratehawk.com"},
            {"name": "_fbp", "value": "fb.1.1713874243068.580299445", "domain": ".ratehawk.com"},
            {"name": "_hcfnl_funnel_uid", "value": "ZielRGYnpUQp33BWZsAFzg==", "domain": ".ratehawk.com"},
            {"name": "_hcfnl_incognito", "value": "0", "domain": ".ratehawk.com"},
            {"name": "__exponea_etc__", "value": "6dd08bc8-5037-4b48-a4dd-a56e52ca1696", "domain": ".ratehawk.com"},
            {"name": "_hcfnl_fpr", "value": "30e32375f722575934b0beb26224493b", "domain": ".ratehawk.com"},
            {"name": "userid", "value": "780476064", "domain": ".ratehawk.com"},
            {"name": "is_auth", "value": "1", "domain": ".ratehawk.com"},
            {"name": "riskified_sid_user", "value": "780476064", "domain": ".ratehawk.com"},
            {"name": "otaSearchFormRooms", "value": "%5B%7B%22adults%22%3A2%2C%22ages%22%3A%5B%5D%7D%5D", "domain": ".ratehawk.com"},
            {"name": "rskxRunCookie", "value": "0", "domain": ".ratehawk.com"},
            {"name": "rCookie", "value": "f9pkt5p6iik7b2u0x5kzmlvccij5d", "domain": ".ratehawk.com"},
            {"name": "otaSelectedCurrencyCode", "value": "USD", "domain": ".ratehawk.com"},
            {"name": "_ym_uid", "value": "1713874945599062555", "domain": ".ratehawk.com"},
            {"name": "_ym_d", "value": "1713874945", "domain": ".ratehawk.com"},
            {"name": "_gcl_aw", "value": "GCL.1713937845.EAIaIQobChMI7o_o_5PahQMVrqaDBx0ssQo6EAAYASAAEgJvFPD_BwE", "domain": ".ratehawk.com"},
            {"name": "_uetvid", "value": "81bfa170016a11efb10173fdfd77bf8b", "domain": ".ratehawk.com"},
            {"name": "_clck", "value": "10sb0qx%7C2%7Cfl7%7C0%7C1574", "domain": ".ratehawk.com"},
            {"name": "sessionid", "value": ".Mz5yLOS6nMgm9MA54WTd2Ke7b0C3jAYhaB3cZREwlzuTFkPSSHVLzm1zcKwXDCNYC80Y9PvFBs75Ilxd_1stG7H1YZmvY9dAfNl_I231kmd-nwsYBB6mZVdYCTBiB3Vx0PiQRKQ3N1Ne16cXekqBkhj6tHthjW5wYRmCCJOBEVC46pB-Fq8Aqnm3dk26x4_EfWK58oxbcgvyjsRQHaJeuQ:1rzVX1:5va7ATYgFJlpWjIV2mz2H9EHgwAhkYdrIB7y8_uh1-c", "domain": ".ratehawk.com"},
            {"name": "_gac_UA-19627229-19", "value": "1.1713937864.EAIaIQobChMI7o_o_5PahQMVrqaDBx0ssQo6EAAYASAAEgJvFPD_BwE", "domain": ".ratehawk.com"},
            {"name": "partner_original_url", "value": "https://www.ratehawk.com/hotel/united_arab_emirates/aqah/mid7492433/miramar_al_aqah_beach_resort/?q=965825047&dates=15.05.2024-16.05.2024&guests=2&residency=en-us&cur=USD&sid=f32a1f64-8055-4769-8dd4-ddd31db80769", "domain": ".ratehawk.com"},
            {"name": "partner_original_referer", "value": "", "domain": ".ratehawk.com"},
            {"name": "_gid", "value": "GA1.2.1921172613.1715754121", "domain": ".ratehawk.com"},
            {"name": "otaSearchFormFrom", "value": "20240813", "domain": ".ratehawk.com"},
            {"name": "otaSearchFormTo", "value": "20240814", "domain": ".ratehawk.com"},
            {"name": "csrftoken", "value": "t8injYevz5cixrGQcznAOsqlAew0ObvfZ5MSx1A2KcWniRhlzXQlnNCcyrJW9PGY", "domain": ".ratehawk.com"},
            {"name": "messages", "value": "", "domain": ".ratehawk.com"},
            {"name": "user_language", "value": "en", "domain": ".ratehawk.com"},
            {"name": "otaPixelRatio", "value": "1.25", "domain": ".ratehawk.com"},
            {"name": "prtnrContractSlug", "value": "68540.b2b.ad98", "domain": ".ratehawk.com"},
            {"name": "userlucky", "value": "62", "domain": ".ratehawk.com"},
            {"name": "_ym_isad", "value": "2", "domain": ".ratehawk.com"},
            {"name": "__exponea_time2__", "value": "-4.158631324768066", "domain": ".ratehawk.com"}
        ]
        for cookie in ratehawk_cookies:
            driver.add_cookie(cookie)
        driver.refresh()

        try:
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'p.zenroomspage-rates-table-cell-room-name-title')))
            print("RateHawk page loaded successfully.")
        except:
            print("RateHawk page did not load correctly.")
            driver.save_screenshot('ratehawk_error.png')
            return []

        ratehawk_room_data = []
        try:
            ratehawk_room_elements = driver.find_elements(By.CSS_SELECTOR, 'p.zenroomspage-rates-table-cell-room-name-title')
            for room_element in ratehawk_room_elements:
                try:
                    room_type = room_element.text.strip()
                except Exception:
                    room_type = "N/A"
                try:
                    meal_plan_element = room_element.find_element(By.XPATH, './/ancestor::tr//div[contains(@class, "valueadds-item-title-inner")]')
                    meal_plan_text = meal_plan_element.text.strip() if meal_plan_element else 'N/A'
                    if 'Meals are not included' in meal_plan_text:
                        meal_plan = 'Room Only'
                    elif 'Breakfast' in meal_plan_text:
                        meal_plan = 'Breakfast'
                    else:
                        meal_plan = meal_plan_text
                except Exception:
                    meal_plan = "N/A"
                try:
                    cancellation_element = room_element.find_element(By.XPATH, './/ancestor::tr//div[contains(@class, "zenroomspage-rates-table-cell-cancellation-penalty-free")]')
                    cancellation_text = cancellation_element.text.strip() if cancellation_element else 'N/A'
                    if cancellation_text == 'N/A':
                        cancellation = 'Non-refundable'
                    else:
                        cancellation = 'Free cancellation'
                except Exception:
                    cancellation = "Non-refundable"
                try:
                    price_element = room_element.find_element(By.XPATH, './/ancestor::tr//div[contains(@class, "zenroomspage-rates-table-cell-price-value-main")]')
                    price_text = price_element.text.strip() if price_element else 'N/A'
                    price = price_text.replace('$', '').replace('\xa0', '').strip()
                except Exception as e:
                    price = f"Error extracting price: {e}"
                
                bedrooms = extract_bedroom_count(room_type)
                
                ratehawk_room_data.append((room_type, meal_plan, cancellation, price, bedrooms))
        except Exception as e:
            print(f"Error during RateHawk scraping: {e}")

        print("Scraping from RateHawk successful.")
        return ratehawk_room_data
    finally:
        driver.quit()
        print("RateHawk driver closed.")

# Function to process data and save to Excel
def process(iol_hotel_url, ratehawk_hotel_url, excel_path, tmpdirname):
    # Scrape data concurrently
    with ThreadPoolExecutor() as executor:
        iol_future = executor.submit(scrape_iol, iol_hotel_url)
        ratehawk_future = executor.submit(scrape_ratehawk, ratehawk_hotel_url)
        
        try:
            iol_room_data, start_date, hotel_name_iol = iol_future.result()
        except Exception as e:
            print(f"IOL scraping failed: {e}")
            iol_room_data, start_date, hotel_name_iol = [], None, None
        
        try:
            ratehawk_room_data = ratehawk_future.result()
        except Exception as e:
            print(f"RateHawk scraping failed: {e}")
            ratehawk_room_data = []

    # Matching and comparison logic
    matched_rooms = []
    for iol_room in iol_room_data:
        iol_room_type = iol_room[0]
        iol_meal_plan = iol_room[1]
        iol_cancellation = iol_room[2]
        iol_bedrooms = iol_room[4]

        matching_rooms = [room for room in ratehawk_room_data if room[4] == iol_bedrooms]

        if matching_rooms:
            best_match = fuzz_process.extractOne(iol_room_type, [r[0] for r in matching_rooms], scorer=fuzz.ratio)
            if best_match:
                match_index = [r[0] for r in matching_rooms].index(best_match[0])
                matched_room = matching_rooms[match_index]
                
                # Compare meal plans and cancellation policy if match scores are equal
                if iol_meal_plan == matched_room[1] and iol_cancellation == matched_room[2]:
                    matched_rooms.append((iol_room, matched_room))
                elif iol_meal_plan == matched_room[1] or iol_cancellation == matched_room[2]:
                    matched_rooms.append((iol_room, matched_room))
                else:
                    best_secondary_match = None
                    for room in matching_rooms:
                        if room[1] == iol_meal_plan and room[2] == iol_cancellation:
                            best_secondary_match = room
                            break
                        elif room[1] == iol_meal_plan or room[2] == iol_cancellation:
                            if not best_secondary_match or (room[1] == iol_meal_plan and room[2] == iol_cancellation):
                                best_secondary_match = room
                    if best_secondary_match:
                        matched_rooms.append((iol_room, best_secondary_match))
                    else:
                        matched_rooms.append((iol_room, matched_room))

    matched_rooms = sorted(matched_rooms, key=lambda x: fuzz_process.extractOne(x[0][0], [r[0] for r in ratehawk_room_data], scorer=fuzz.ratio)[1], reverse=True)[:3]

    for i, (iol_room, matched_room) in enumerate(matched_rooms, start=1):
        print(f"Matched Room {i}:")
        print(f"IOL - Room Type: {iol_room[0]}, Meal Plan: {iol_room[1]}, Cancellation: {iol_room[2]}, Price: {iol_room[3]}")
        print(f"RateHawk - Room Type: {matched_room[0]}, Meal Plan: {matched_room[1]}, Cancellation: {matched_room[2]}, Price: {matched_room[3]}")

    print("Comparison complete.")

    # Write results to Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    def find_next_row(ws, column):
        for row in range(2, ws.max_row + 2):
            cell = ws.cell(row=row, column=column)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell) and cell.value is None:
                return row
        return ws.max_row + 1

    next_row = find_next_row(ws, 2)
    ws[f'B{next_row}'] = hotel_name_iol
    print(f"Written hotel name {hotel_name_iol} to column B{next_row}")

    for i, (iol_room, matched_room) in enumerate(matched_rooms):
        row = next_row + i
        iol_room_type, iol_meal_plan, iol_cancellation, iol_price, _ = iol_room
        ratehawk_room_type, ratehawk_meal_plan, ratehawk_cancellation, ratehawk_price, _ = matched_room

        ws.cell(row=row, column=7).value = iol_room_type
        print(f"Written IOL room type {iol_room_type} to column G{row}")
        ws.cell(row=row, column=10).value = iol_meal_plan
        print(f"Written IOL meal plan {iol_meal_plan} to column J{row}")
        ws.cell(row=row, column=8).value = start_date  # Place the date in the desired column
        print(f"Written start date {start_date} to column K{row}")
        ws.cell(row=row, column=12).value = iol_price
        print(f"Written IOL price {iol_price} to column L{row}")
        ws.cell(row=row, column=16).value = ratehawk_price
        print(f"Written RateHawk price {ratehawk_price} to column P{row}")

    # Save the result in a consistent format to the temporary directory
    result_file_path = os.path.join(tmpdirname, 'processed_' + os.path.basename(excel_path))
    print("Saving Excel file to:", result_file_path)
    wb.save(result_file_path)
    wb.close()  # Ensure the workbook is closed properly
    print("Data written to Excel file.")
    
    return result_file_path
